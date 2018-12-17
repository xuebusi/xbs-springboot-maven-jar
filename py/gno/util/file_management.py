import csv
from os import listdir, path

import numpy as np
import openpyxl


def get_filenames(path_to_files, extension):
    image_filenames = []
    for file in listdir(path_to_files):
        if file.endswith('.' + extension):
            image_filenames = image_filenames + [file]

    return image_filenames


def read_csv_classification_results(csv_filename):
    image_filenames = []
    scores = []

    with open(csv_filename, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader)
        for row in csv_reader:
            image_filenames = image_filenames + [row[0]]
            scores = scores + [float(row[1])]

    scores = np.asarray(scores, dtype=np.float)

    return image_filenames, scores


def get_labels_from_training_data(gt_folder):
    glaucoma_folder = path.join(gt_folder, 'Glaucoma')
    non_glaucoma_folder = path.join(gt_folder, 'Non-Glaucoma')

    glaucoma_filenames = get_filenames(glaucoma_folder, 'bmp')
    non_glaucoma_filenames = get_filenames(non_glaucoma_folder, 'bmp')

    image_filenames = glaucoma_filenames + non_glaucoma_filenames

    labels = np.zeros(len(image_filenames), dtype=np.bool)
    labels[0:len(glaucoma_filenames)] = True

    return image_filenames, labels


def read_gt_labels(xlsx_filename):
    image_filenames = []
    labels = None

    book = openpyxl.load_workbook(xlsx_filename)
    current_sheet = book.active

    for row in current_sheet.iter_rows(min_row=2, min_col=1):
        current_name = row[0].value[:-3] + 'jpg'
        image_filenames = image_filenames + [current_name]
        current_label = row[1].value > 0
        if labels is None:
            labels = current_label
        else:
            labels = np.vstack((labels, current_label))

    return image_filenames, labels


def sort_scores_by_filename(target_names, names_to_sort, values_to_sort):
    names_to_sort = [x.upper() for x in names_to_sort]

    sorted_values = np.zeros(values_to_sort.shape)

    for i in range(len(target_names)):
        sorted_values[i] = values_to_sort[names_to_sort.index(target_names[i].upper())]

    return sorted_values


def read_fovea_location_results(csv_filename):
    image_filenames = []
    coordinates = None

    with open(csv_filename, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        next(csv_reader)
        for row in csv_reader:
            image_filenames = image_filenames + [row[0]]
            current_coordinates = np.asarray(row[1:], dtype=np.float)
            if coordinates is None:
                coordinates = current_coordinates
            else:
                coordinates = np.vstack((coordinates, current_coordinates))

    return image_filenames, coordinates


def read_gt_fovea_location(xlsx_filename, is_training=False):
    image_filenames = []
    coordinates = None

    book = openpyxl.load_workbook(xlsx_filename)
    current_sheet = book.active

    for row in current_sheet.iter_rows(min_row=2, min_col=1):
        image_filenames = image_filenames + [row[1].value]
        if is_training:
            current_coordinates = np.asarray([float(row[2].value), float(row[3].value)], dtype=np.float)
        else:
            current_coordinates = np.asarray([float(row[3].value), float(row[4].value)], dtype=np.float)
        if coordinates is None:
            coordinates = current_coordinates
        else:
            coordinates = np.vstack((coordinates, current_coordinates))

    return image_filenames, coordinates


def sort_coordinates_by_filename(target_names, names_to_sort, values_to_sort):
    sorted_values = np.zeros(values_to_sort.shape)

    for i in range(len(target_names)):
        sorted_values[i, :] = values_to_sort[names_to_sort.index(target_names[i])]

    return sorted_values
